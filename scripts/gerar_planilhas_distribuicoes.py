"""Gera planilhas de pratica para a Aula 2: Estatistica para RH.

Cria dois arquivos em ``dados/``:

- ``distribuicoes_exercicio.xlsx``: versao em branco, com a base de dados,
  uma aba de instrucoes e abas vazias para a pessoa estudante preencher.
- ``distribuicoes_resolvido.xlsx``: mesma base, mas com abas resolvidas
  contendo estatisticas descritivas, testes de normalidade, histogramas,
  arvore de decisao para escolha do teste, e testes parametricos e nao
  parametricos com interpretacao.

O objetivo didatico e ligar tres ideias que costumam aparecer soltas:

1. Como os dados se distribuem (forma, simetria, cauda).
2. Quais sao as suposicoes de cada teste estatistico.
3. Como escolher entre um teste parametrico e um nao parametrico.

A base e sintetica e reproduzivel (seed fixa). Usa colunas com
distribuicoes propositalmente diferentes para evidenciar como uma mesma
pergunta exige testes diferentes dependendo do formato dos dados.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from scipy import stats

ROOT = Path(__file__).resolve().parents[1]
DADOS = ROOT / "dados"
SEED = 42
N = 1200  # quantidade de observacoes por coluna

# ---------------------------------------------------------------------------
# Estilos
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SECTION_FONT = Font(bold=True, size=12, color="1F4E78")
INTERP_FONT = Font(italic=True, color="595959")
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
RED_FILL = PatternFill("solid", fgColor="F8CBAD")
YELLOW_FILL = PatternFill("solid", fgColor="FFE699")
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)

NUMERIC_COLS = [
    "salario_mensal",
    "idade",
    "tempo_empresa_meses",
    "horas_treinamento_ano",
    "nota_avaliacao",
    "dias_absenteismo_ano",
    "score_engajamento",
    "horas_extras_mes",
]


# ---------------------------------------------------------------------------
# Geracao da base
# ---------------------------------------------------------------------------


@dataclass
class GeneratedData:
    df: pd.DataFrame
    descricoes: dict[str, str]


def gerar_base(n: int = N, seed: int = SEED) -> GeneratedData:
    """Gera ``n`` observacoes sinteticas com colunas de distribuicoes variadas.

    A escolha das distribuicoes e proposital:

    - ``salario_mensal``: lognormal -> assimetrica a direita, cauda longa.
    - ``idade``: normal truncada -> simetrica, sino classico.
    - ``tempo_empresa_meses``: exponencial -> muito assimetrica.
    - ``horas_treinamento_ano``: gamma -> assimetrica leve.
    - ``nota_avaliacao``: normal restrita a 1..5 -> aproximadamente simetrica.
    - ``dias_absenteismo_ano``: Poisson -> contagem, assimetrica.
    - ``score_engajamento``: normal -> simetrica.
    - ``horas_extras_mes``: mistura bimodal -> nao normal, bimodal.
    """
    rng = np.random.default_rng(seed)

    salario = rng.lognormal(mean=np.log(4500), sigma=0.55, size=n).round(2)
    idade = np.clip(rng.normal(loc=37, scale=9, size=n), 18, 65).round(0).astype(int)
    tempo_empresa = rng.exponential(scale=42, size=n).round(0).astype(int)
    horas_treinamento = rng.gamma(shape=2.0, scale=12, size=n).round(1)
    nota_avaliacao = np.clip(rng.normal(loc=3.5, scale=0.6, size=n), 1, 5).round(2)
    dias_abs = rng.poisson(lam=4.5, size=n)
    engajamento = np.clip(rng.normal(loc=70, scale=12, size=n), 0, 100).round(1)

    # Bimodal: alguns colaboradores quase nao fazem hora extra, outros fazem muito
    pico_baixo = rng.normal(loc=4, scale=2, size=n)
    pico_alto = rng.normal(loc=28, scale=6, size=n)
    mascara_alto = rng.random(size=n) < 0.35
    horas_extras = np.where(mascara_alto, pico_alto, pico_baixo)
    horas_extras = np.clip(horas_extras, 0, None).round(1)

    area = rng.choice(
        ["Operacoes", "Comercial", "TI", "Administrativo"],
        size=n,
        p=[0.45, 0.25, 0.15, 0.15],
    )
    genero = rng.choice(["F", "M"], size=n, p=[0.48, 0.52])
    tipo_contrato = rng.choice(["CLT", "PJ"], size=n, p=[0.82, 0.18])

    df = pd.DataFrame(
        {
            "colaborador_id": np.arange(1, n + 1),
            "area": area,
            "genero": genero,
            "tipo_contrato": tipo_contrato,
            "salario_mensal": salario,
            "idade": idade,
            "tempo_empresa_meses": tempo_empresa,
            "horas_treinamento_ano": horas_treinamento,
            "nota_avaliacao": nota_avaliacao,
            "dias_absenteismo_ano": dias_abs,
            "score_engajamento": engajamento,
            "horas_extras_mes": horas_extras,
        }
    )

    descricoes = {
        "salario_mensal": "Lognormal: assimetrica a direita, cauda longa (tipico de salarios).",
        "idade": "Normal truncada entre 18 e 65: simetrica, formato de sino.",
        "tempo_empresa_meses": "Exponencial: muita gente com pouco tempo, poucos com muito tempo.",
        "horas_treinamento_ano": "Gamma: assimetria moderada a direita.",
        "nota_avaliacao": "Normal limitada entre 1 e 5: quase simetrica.",
        "dias_absenteismo_ano": "Poisson com lambda 4,5: contagem assimetrica.",
        "score_engajamento": "Normal entre 0 e 100: simetrica.",
        "horas_extras_mes": "Bimodal: mistura de quem faz pouca e quem faz muita hora extra.",
    }
    return GeneratedData(df=df, descricoes=descricoes)


# ---------------------------------------------------------------------------
# Helpers de escrita
# ---------------------------------------------------------------------------


def escrever_dataframe(ws, df: pd.DataFrame, start_row: int = 1, start_col: int = 1) -> None:
    """Escreve um DataFrame em ``ws`` a partir de (start_row, start_col)."""
    for r_offset, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
        for c_offset, value in enumerate(row):
            cell = ws.cell(row=start_row + r_offset, column=start_col + c_offset, value=value)
            if r_offset == 0:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT


def autoajustar(ws, min_width: int = 10, max_width: int = 32) -> None:
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        largura = min_width
        for cell in col_cells:
            valor = cell.value
            if valor is None:
                continue
            largura = max(largura, min(max_width, len(str(valor)) + 2))
        ws.column_dimensions[letter].width = largura


def escrever_paragrafo(ws, linhas: Iterable[str], start_row: int, col: int = 1) -> int:
    """Escreve cada string como uma linha; retorna a proxima linha livre."""
    r = start_row
    for linha in linhas:
        cell = ws.cell(row=r, column=col, value=linha)
        cell.alignment = LEFT
        r += 1
    return r


# ---------------------------------------------------------------------------
# Abas
# ---------------------------------------------------------------------------


def aba_instrucoes(wb: Workbook, resolvido: bool) -> None:
    ws = wb.create_sheet("Instrucoes", 0)
    titulo = ws.cell(row=1, column=1, value="Pratica guiada: distribuicoes e escolha do teste estatistico")
    titulo.font = Font(bold=True, size=14, color="1F4E78")

    linhas = [
        "",
        "Esta planilha acompanha a Aula 2 do curso. A ideia e treinar o olho",
        "para reconhecer formatos de distribuicao e ligar isso a escolha do teste.",
        "",
        "Roteiro sugerido:",
        "1) Aba 'Dados': observe as 1.200 linhas. Cada coluna numerica tem uma",
        "   distribuicao diferente de proposito.",
        "2) Aba 'Estatisticas': calcule (ou confira) media, mediana, desvio,",
        "   minimo, maximo e quartis para cada coluna numerica.",
        "3) Aba 'Histogramas': plote o histograma de cada variavel numerica.",
        "   Pergunte-se: o formato lembra um sino simetrico? Tem cauda longa?",
        "   Tem dois picos? Tem muitos zeros?",
        "4) Aba 'Normalidade': aplique um teste de normalidade (Shapiro-Wilk",
        "   ate ~5.000 observacoes, ou Kolmogorov-Smirnov). Lembre que com",
        "   amostras grandes quase tudo 'rejeita' normalidade; combine o teste",
        "   com a inspecao visual.",
        "5) Aba 'Decisao': para cada pergunta de RH, decida se vai usar um",
        "   teste parametrico (assume normalidade e variancias parecidas) ou",
        "   um teste nao parametrico (baseado em postos, sem essa exigencia).",
        "6) Abas de testes: rode o teste escolhido e escreva uma interpretacao",
        "   em portugues claro, sem jargao.",
        "",
        "Parametricos x nao parametricos em uma frase:",
        "- Parametricos confiam que os dados seguem uma forma conhecida",
        "  (normal, geralmente) e por isso usam media e desvio.",
        "- Nao parametricos nao exigem esse formato; usam postos (ranks) e",
        "  por isso sao mais robustos contra outliers e assimetria.",
        "",
        "Regra pratica para escolher:",
        "- Amostra pequena (n < 30) e distribuicao claramente nao normal -> nao parametrico.",
        "- Amostra grande e distribuicao razoavelmente simetrica -> parametrico costuma servir.",
        "- Outliers extremos ou cauda muito longa -> considere nao parametrico ou transforme",
        "  os dados (por exemplo, log de salario).",
        "- Variavel categorica versus categorica -> qui-quadrado.",
        "- Duas variaveis numericas -> correlacao de Pearson (linear) ou Spearman (postos).",
    ]
    if resolvido:
        linhas += [
            "",
            "Esta e a versao RESOLVIDA: as abas a seguir ja trazem calculos,",
            "graficos e interpretacoes. Use como gabarito depois de tentar.",
        ]
    else:
        linhas += [
            "",
            "Esta e a versao em BRANCO: abas vazias estao preparadas com",
            "cabecalhos para voce preencher. Use as formulas do Excel:",
            "- MEDIA, MED, DESVPAD.A, MINIMO, MAXIMO, QUARTIL.INC",
            "- FREQUENCIA + grafico de colunas para histograma",
            "- Para Shapiro-Wilk e Mann-Whitney, faça em Python/R (a versao",
            "  resolvida traz os resultados ja calculados pelo scipy).",
        ]
    escrever_paragrafo(ws, linhas, start_row=2)
    ws.column_dimensions["A"].width = 95


def aba_dados(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Dados")
    escrever_dataframe(ws, df)
    ws.freeze_panes = "A2"
    autoajustar(ws)


def aba_dicionario(wb: Workbook, descricoes: dict[str, str]) -> None:
    ws = wb.create_sheet("Dicionario")
    df = pd.DataFrame(
        [{"coluna": k, "distribuicao_esperada": v} for k, v in descricoes.items()]
    )
    escrever_dataframe(ws, df)
    autoajustar(ws, max_width=70)


def aba_estatisticas(wb: Workbook, df: pd.DataFrame, preencher: bool) -> None:
    ws = wb.create_sheet("Estatisticas")
    cabecalho = [
        "coluna",
        "n",
        "media",
        "mediana",
        "desvio_padrao",
        "minimo",
        "Q1",
        "Q3",
        "maximo",
        "assimetria_skew",
        "curtose",
    ]
    for c, nome in enumerate(cabecalho, start=1):
        cell = ws.cell(row=1, column=c, value=nome)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for r, col in enumerate(NUMERIC_COLS, start=2):
        ws.cell(row=r, column=1, value=col)
        if preencher:
            s = df[col]
            ws.cell(row=r, column=2, value=int(s.count()))
            ws.cell(row=r, column=3, value=round(float(s.mean()), 4))
            ws.cell(row=r, column=4, value=round(float(s.median()), 4))
            ws.cell(row=r, column=5, value=round(float(s.std(ddof=1)), 4))
            ws.cell(row=r, column=6, value=round(float(s.min()), 4))
            ws.cell(row=r, column=7, value=round(float(s.quantile(0.25)), 4))
            ws.cell(row=r, column=8, value=round(float(s.quantile(0.75)), 4))
            ws.cell(row=r, column=9, value=round(float(s.max()), 4))
            ws.cell(row=r, column=10, value=round(float(stats.skew(s)), 4))
            ws.cell(row=r, column=11, value=round(float(stats.kurtosis(s)), 4))

    if preencher:
        r = len(NUMERIC_COLS) + 4
        ws.cell(row=r, column=1, value="Como ler:").font = SECTION_FONT
        leitura = [
            "- Se media e mediana sao muito diferentes, a distribuicao e assimetrica.",
            "- Assimetria (skew) proxima de 0 indica simetria. Positiva = cauda a direita.",
            "- Curtose alta indica cauda mais pesada que a normal.",
            "- Desvio padrao grande em relacao a media indica muita dispersao.",
        ]
        escrever_paragrafo(ws, leitura, start_row=r + 1)
    else:
        r = len(NUMERIC_COLS) + 4
        ws.cell(row=r, column=1, value="Dica:").font = SECTION_FONT
        dica = [
            "Use as formulas do Excel apontando para a aba 'Dados':",
            "=MEDIA(Dados!E2:E1201), =MED(Dados!E2:E1201), =DESVPAD.A(...),",
            "=QUARTIL.INC(intervalo;1) e =QUARTIL.INC(intervalo;3).",
            "Para skewness use =DISTORCAO(...) e para curtose =CURT(...).",
        ]
        escrever_paragrafo(ws, dica, start_row=r + 1)

    autoajustar(ws)


def aba_histogramas(wb: Workbook, df: pd.DataFrame, preencher: bool) -> None:
    ws = wb.create_sheet("Histogramas")
    ws.cell(row=1, column=1, value="Histogramas das variaveis numericas").font = SECTION_FONT
    ws.cell(
        row=2,
        column=1,
        value=(
            "Cada bloco abaixo mostra os limites do bin (faixa) e a frequencia. "
            "O grafico ao lado traduz isso em um histograma. Procure: simetria, "
            "cauda, picos multiplos."
        ),
    ).alignment = LEFT
    ws.row_dimensions[2].height = 32

    linha_atual = 4
    for col in NUMERIC_COLS:
        ws.cell(row=linha_atual, column=1, value=col).font = SECTION_FONT
        ws.cell(row=linha_atual, column=2, value=f"({len(df)} obs)").font = INTERP_FONT
        linha_atual += 1

        ws.cell(row=linha_atual, column=1, value="bin_centro")
        ws.cell(row=linha_atual, column=2, value="frequencia")
        for c in (1, 2):
            ws.cell(row=linha_atual, column=c).fill = HEADER_FILL
            ws.cell(row=linha_atual, column=c).font = HEADER_FONT
        cabecalho_row = linha_atual
        linha_atual += 1

        if preencher:
            valores = df[col].dropna().to_numpy()
            counts, edges = np.histogram(valores, bins=25)
            centros = (edges[:-1] + edges[1:]) / 2
            for centro, qtd in zip(centros, counts):
                ws.cell(row=linha_atual, column=1, value=round(float(centro), 3))
                ws.cell(row=linha_atual, column=2, value=int(qtd))
                linha_atual += 1

            chart = BarChart()
            chart.type = "col"
            chart.style = 11
            chart.title = f"Histograma: {col}"
            chart.y_axis.title = "Frequencia"
            chart.x_axis.title = col
            chart.legend = None
            chart.height = 8
            chart.width = 16
            data_ref = Reference(
                ws,
                min_col=2,
                min_row=cabecalho_row,
                max_row=cabecalho_row + len(counts),
                max_col=2,
            )
            cat_ref = Reference(
                ws,
                min_col=1,
                min_row=cabecalho_row + 1,
                max_row=cabecalho_row + len(counts),
            )
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cat_ref)
            ws.add_chart(chart, f"D{cabecalho_row}")
            linha_atual += 2  # espaco entre blocos
        else:
            ws.cell(
                row=linha_atual,
                column=1,
                value=(
                    "Preencha aqui os bins (use =FREQUENCIA) e gere um grafico"
                    " de colunas usando bin_centro x frequencia."
                ),
            ).alignment = LEFT
            linha_atual += 28  # espaco vazio para o estudante

    autoajustar(ws, max_width=22)


def _classificar_normalidade(p: float) -> tuple[str, PatternFill]:
    if p < 0.01:
        return ("Rejeita normalidade (p < 0,01).", RED_FILL)
    if p < 0.05:
        return ("Provavelmente nao normal (p < 0,05).", RED_FILL)
    if p < 0.10:
        return ("Zona cinza (0,05 <= p < 0,10).", YELLOW_FILL)
    return ("Nao rejeita normalidade (p >= 0,10).", GREEN_FILL)


def aba_normalidade(wb: Workbook, df: pd.DataFrame, preencher: bool) -> None:
    ws = wb.create_sheet("Normalidade")
    explicacao = [
        "Teste de normalidade: avalia se a distribuicao dos dados pode ser",
        "considerada compativel com uma distribuicao normal.",
        "",
        "Testes usados aqui:",
        "- Shapiro-Wilk: bom para amostras pequenas e medias (ate ~5.000).",
        "- D'Agostino-Pearson: combina assimetria e curtose; bom para n grande.",
        "",
        "Atencao: com amostras grandes (n=1.200, por exemplo), quase qualquer",
        "desvio pequeno faz o teste rejeitar normalidade. Por isso, combine",
        "o p-valor com o histograma e com a comparacao media x mediana.",
    ]
    escrever_paragrafo(ws, explicacao, start_row=1)
    ws.column_dimensions["A"].width = 30

    cabecalho = [
        "coluna",
        "n",
        "Shapiro_W",
        "Shapiro_p",
        "DAgostino_stat",
        "DAgostino_p",
        "interpretacao",
        "sugestao_de_teste",
    ]
    header_row = len(explicacao) + 3
    for c, nome in enumerate(cabecalho, start=1):
        cell = ws.cell(row=header_row, column=c, value=nome)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    if preencher:
        sugestoes = {
            "salario_mensal": "Mann-Whitney / Spearman; ou log-transformar e usar parametrico.",
            "idade": "Pode usar teste t e Pearson (suficientemente simetrica).",
            "tempo_empresa_meses": "Mann-Whitney / Kruskal-Wallis (muito assimetrica).",
            "horas_treinamento_ano": "Mann-Whitney ou transformar (raiz, log).",
            "nota_avaliacao": "Teste t / ANOVA (quase simetrica, escala limitada).",
            "dias_absenteismo_ano": "Mann-Whitney / Kruskal-Wallis (contagem com cauda).",
            "score_engajamento": "Teste t / ANOVA (simetrica).",
            "horas_extras_mes": "Nao parametrico: distribuicao bimodal viola normalidade.",
        }
        for r, col in enumerate(NUMERIC_COLS, start=header_row + 1):
            s = df[col].dropna()
            amostra = s.sample(n=min(5000, len(s)), random_state=SEED)
            sw_stat, sw_p = stats.shapiro(amostra)
            dap_stat, dap_p = stats.normaltest(s)
            interpretacao, fill = _classificar_normalidade(min(sw_p, dap_p))

            ws.cell(row=r, column=1, value=col)
            ws.cell(row=r, column=2, value=int(s.count()))
            ws.cell(row=r, column=3, value=round(float(sw_stat), 4))
            ws.cell(row=r, column=4, value=float(f"{sw_p:.4g}"))
            ws.cell(row=r, column=5, value=round(float(dap_stat), 4))
            ws.cell(row=r, column=6, value=float(f"{dap_p:.4g}"))
            cell_int = ws.cell(row=r, column=7, value=interpretacao)
            cell_int.fill = fill
            ws.cell(row=r, column=8, value=sugestoes[col])
    else:
        for r, col in enumerate(NUMERIC_COLS, start=header_row + 1):
            ws.cell(row=r, column=1, value=col)

    autoajustar(ws, max_width=55)


def aba_decisao(wb: Workbook, preencher: bool) -> None:
    ws = wb.create_sheet("Decisao")
    titulo = ws.cell(row=1, column=1, value="Arvore de decisao para escolher o teste")
    titulo.font = SECTION_FONT

    arvore = [
        "",
        "1) Qual e o tipo de cada variavel envolvida?",
        "   - Numerica continua",
        "   - Numerica de contagem",
        "   - Categorica (duas ou mais categorias)",
        "",
        "2) Quantos grupos voce esta comparando?",
        "   - 1 grupo contra um valor de referencia -> teste t de uma amostra",
        "     (parametrico) ou teste do sinal/Wilcoxon (nao parametrico)",
        "   - 2 grupos independentes -> teste t independente OU Mann-Whitney",
        "   - 2 grupos pareados -> teste t pareado OU Wilcoxon pareado",
        "   - 3 ou mais grupos -> ANOVA OU Kruskal-Wallis",
        "",
        "3) Os dados respeitam as suposicoes do teste parametrico?",
        "   - Distribuicao aproximadamente normal em cada grupo?",
        "   - Variancias parecidas entre os grupos? (teste de Levene)",
        "   - Observacoes independentes?",
        "   Se SIM nas tres -> use o teste parametrico (mais potente).",
        "   Se NAO -> prefira o nao parametrico (mais robusto).",
        "",
        "4) Para duas variaveis numericas:",
        "   - Pearson assume relacao linear e normalidade aproximada.",
        "   - Spearman usa postos, captura relacoes monotonicas sem exigir normalidade.",
        "",
        "5) Para duas variaveis categoricas:",
        "   - Qui-quadrado de independencia (se as contagens esperadas forem >= 5).",
        "   - Se contagens muito pequenas, use teste exato de Fisher.",
    ]
    escrever_paragrafo(ws, arvore, start_row=2)

    if preencher:
        ws.cell(row=len(arvore) + 4, column=1, value="Aplicando ao nosso dataset:").font = SECTION_FONT
        aplicacoes = [
            (
                "Comparar salario_mensal entre genero F e M",
                "salario nao e normal (lognormal). Use Mann-Whitney. Para parametrico, log-transformar e teste t.",
            ),
            (
                "Comparar score_engajamento entre as 4 areas",
                "score e aproximadamente normal. Use ANOVA. Confirme variancias com Levene.",
            ),
            (
                "Comparar dias_absenteismo_ano entre areas",
                "contagem assimetrica. Use Kruskal-Wallis (nao parametrico).",
            ),
            (
                "Associacao entre area e tipo_contrato",
                "duas categoricas. Use qui-quadrado de independencia.",
            ),
            (
                "Relacao entre horas_treinamento_ano e nota_avaliacao",
                "uma assimetrica e uma simetrica. Use Spearman; Pearson como comparacao.",
            ),
        ]
        df_apl = pd.DataFrame(aplicacoes, columns=["pergunta_de_RH", "teste_indicado_e_motivo"])
        escrever_dataframe(ws, df_apl, start_row=len(arvore) + 5)

    autoajustar(ws, max_width=70)


def aba_teste_t(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Teste_t_e_ANOVA")
    ws.cell(row=1, column=1, value="Testes parametricos com interpretacao").font = SECTION_FONT

    # Teste t: score_engajamento entre F e M
    grupo_f = df.loc[df["genero"] == "F", "score_engajamento"]
    grupo_m = df.loc[df["genero"] == "M", "score_engajamento"]
    t_stat, t_p = stats.ttest_ind(grupo_f, grupo_m, equal_var=False)
    lev_stat, lev_p = stats.levene(grupo_f, grupo_m)

    linhas_t = [
        "",
        "1) Teste t de Welch: score_engajamento entre genero F e M",
        f"   Media F = {grupo_f.mean():.2f} (n={len(grupo_f)})",
        f"   Media M = {grupo_m.mean():.2f} (n={len(grupo_m)})",
        f"   Levene (variancias iguais): stat={lev_stat:.3f}, p={lev_p:.4g}",
        f"   t = {t_stat:.3f}, p = {t_p:.4g}",
        (
            "   Interpretacao: se p < 0,05, ha evidencia de diferenca media. "
            "Se p >= 0,05, nao temos evidencia suficiente. Sempre relate o tamanho do efeito."
        ),
    ]
    escrever_paragrafo(ws, linhas_t, start_row=2)

    # ANOVA: score_engajamento entre as 4 areas
    grupos_area = [df.loc[df["area"] == a, "score_engajamento"].to_numpy() for a in df["area"].unique()]
    f_stat, f_p = stats.f_oneway(*grupos_area)
    medias_area = df.groupby("area")["score_engajamento"].agg(["count", "mean", "std"]).round(2)

    proxima = len(linhas_t) + 4
    ws.cell(row=proxima, column=1, value="2) ANOVA: score_engajamento entre areas").font = SECTION_FONT
    medias_area_df = medias_area.reset_index()
    escrever_dataframe(ws, medias_area_df, start_row=proxima + 1)
    proxima2 = proxima + 2 + len(medias_area_df)
    ws.cell(row=proxima2, column=1, value=f"F = {f_stat:.3f}, p = {f_p:.4g}")
    ws.cell(
        row=proxima2 + 1,
        column=1,
        value=(
            "Interpretacao: a ANOVA diz se existe ao menos uma media diferente. "
            "Para saber quais grupos diferem, faca um post-hoc (Tukey HSD)."
        ),
    ).alignment = LEFT

    autoajustar(ws, max_width=60)


def aba_nao_parametrico(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("NaoParametricos")
    ws.cell(row=1, column=1, value="Testes nao parametricos com interpretacao").font = SECTION_FONT

    f = df.loc[df["genero"] == "F", "salario_mensal"]
    m = df.loc[df["genero"] == "M", "salario_mensal"]
    u_stat, u_p = stats.mannwhitneyu(f, m, alternative="two-sided")

    linhas_mw = [
        "",
        "1) Mann-Whitney: salario_mensal entre genero F e M",
        f"   Mediana F = {f.median():.2f}",
        f"   Mediana M = {m.median():.2f}",
        f"   U = {u_stat:.0f}, p = {u_p:.4g}",
        (
            "   Quando usar: salario e lognormal, viola normalidade. "
            "Mann-Whitney compara distribuicoes inteiras usando postos."
        ),
    ]
    escrever_paragrafo(ws, linhas_mw, start_row=2)

    grupos = [df.loc[df["area"] == a, "dias_absenteismo_ano"].to_numpy() for a in df["area"].unique()]
    h_stat, h_p = stats.kruskal(*grupos)
    medianas = df.groupby("area")["dias_absenteismo_ano"].median().round(2).reset_index()

    proxima = len(linhas_mw) + 4
    ws.cell(
        row=proxima, column=1, value="2) Kruskal-Wallis: dias_absenteismo_ano entre areas"
    ).font = SECTION_FONT
    escrever_dataframe(ws, medianas, start_row=proxima + 1)
    proxima2 = proxima + 2 + len(medianas)
    ws.cell(row=proxima2, column=1, value=f"H = {h_stat:.3f}, p = {h_p:.4g}")
    ws.cell(
        row=proxima2 + 1,
        column=1,
        value=(
            "Quando usar: contagem assimetrica (Poisson) e 3+ grupos. Kruskal-Wallis "
            "e o equivalente nao parametrico da ANOVA."
        ),
    ).alignment = LEFT

    autoajustar(ws, max_width=60)


def aba_qui_quadrado(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("QuiQuadrado")
    ws.cell(row=1, column=1, value="Qui-quadrado: area x tipo_contrato").font = SECTION_FONT

    tabela = pd.crosstab(df["area"], df["tipo_contrato"])
    chi2, p, gl, esperado = stats.chi2_contingency(tabela)

    ws.cell(row=3, column=1, value="Tabela observada").font = Font(bold=True)
    escrever_dataframe(ws, tabela.reset_index(), start_row=4)

    linha = 4 + len(tabela) + 3
    ws.cell(row=linha, column=1, value="Tabela esperada (sob H0 de independencia)").font = Font(bold=True)
    esperado_df = pd.DataFrame(esperado, index=tabela.index, columns=tabela.columns).round(1)
    escrever_dataframe(ws, esperado_df.reset_index(), start_row=linha + 1)

    linha2 = linha + 1 + len(esperado_df) + 2
    ws.cell(row=linha2, column=1, value=f"Chi2 = {chi2:.3f}, gl = {gl}, p = {p:.4g}")
    ws.cell(
        row=linha2 + 1,
        column=1,
        value=(
            "Interpretacao: H0 = area e tipo de contrato sao independentes. "
            "Se p < 0,05, ha associacao. Olhe a tabela esperada x observada para entender onde."
        ),
    ).alignment = LEFT

    autoajustar(ws)


def aba_correlacao(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Correlacao")
    ws.cell(row=1, column=1, value="Correlacao: Pearson vs Spearman").font = SECTION_FONT
    ws.cell(
        row=2,
        column=1,
        value=(
            "Pearson mede associacao linear e e sensivel a outliers. "
            "Spearman usa postos: captura relacoes monotonicas e e robusto."
        ),
    ).alignment = LEFT

    pearson = df[NUMERIC_COLS].corr(method="pearson").round(3)
    spearman = df[NUMERIC_COLS].corr(method="spearman").round(3)

    ws.cell(row=4, column=1, value="Matriz de Pearson").font = Font(bold=True)
    escrever_dataframe(ws, pearson.reset_index(), start_row=5)

    linha = 5 + len(pearson) + 2
    ws.cell(row=linha, column=1, value="Matriz de Spearman").font = Font(bold=True)
    escrever_dataframe(ws, spearman.reset_index(), start_row=linha + 1)

    linha2 = linha + 1 + len(spearman) + 2
    ws.cell(
        row=linha2,
        column=1,
        value=(
            "Como ler: valores proximos de 1 ou -1 indicam associacao forte. "
            "Diferencas entre Pearson e Spearman sugerem relacao nao linear ou outliers."
        ),
    ).alignment = LEFT

    autoajustar(ws)


# ---------------------------------------------------------------------------
# Construcao dos workbooks
# ---------------------------------------------------------------------------


def construir_workbook(dados: GeneratedData, resolvido: bool) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    aba_instrucoes(wb, resolvido=resolvido)
    aba_dados(wb, dados.df)
    aba_dicionario(wb, dados.descricoes)
    aba_estatisticas(wb, dados.df, preencher=resolvido)
    aba_histogramas(wb, dados.df, preencher=resolvido)
    aba_normalidade(wb, dados.df, preencher=resolvido)
    aba_decisao(wb, preencher=resolvido)

    if resolvido:
        aba_teste_t(wb, dados.df)
        aba_nao_parametrico(wb, dados.df)
        aba_qui_quadrado(wb, dados.df)
        aba_correlacao(wb, dados.df)
    else:
        # Esqueleto vazio para o estudante
        for nome, titulo in [
            ("Teste_t_e_ANOVA", "Espaco para rodar teste t e ANOVA"),
            ("NaoParametricos", "Espaco para Mann-Whitney e Kruskal-Wallis"),
            ("QuiQuadrado", "Espaco para qui-quadrado area x tipo_contrato"),
            ("Correlacao", "Espaco para matriz de correlacao Pearson e Spearman"),
        ]:
            ws = wb.create_sheet(nome)
            ws.cell(row=1, column=1, value=titulo).font = SECTION_FONT
            ws.cell(
                row=3,
                column=1,
                value="Calcule em Python/R e cole os resultados aqui, com interpretacao em portugues claro.",
            ).alignment = LEFT
            ws.column_dimensions["A"].width = 80

    return wb


def main() -> None:
    DADOS.mkdir(parents=True, exist_ok=True)
    dados = gerar_base()

    wb_exercicio = construir_workbook(dados, resolvido=False)
    wb_resolvido = construir_workbook(dados, resolvido=True)

    caminho_ex = DADOS / "distribuicoes_exercicio.xlsx"
    caminho_re = DADOS / "distribuicoes_resolvido.xlsx"
    wb_exercicio.save(caminho_ex)
    wb_resolvido.save(caminho_re)

    print(f"Gerado: {caminho_ex}")
    print(f"Gerado: {caminho_re}")


if __name__ == "__main__":
    main()
